#!/usr/bin/env python3
"""Validate and analyse the final news-source-strategies study.

The unit of replication is one workbook/seed. Contrasts are paired by seed so
that common random numbers reduce Monte Carlo noise.
"""

from __future__ import annotations

import argparse
from concurrent.futures import ProcessPoolExecutor
import hashlib
import json
import math
from pathlib import Path
import warnings

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import pandas as pd
from scipy import stats


SOURCES = ["TRADITIONAL_MEDIA", "UNKNOWN_MEDIA", "FAKE_NEWS_SOURCE", "MIXED_SOURCE"]
TARGET = "FAKE_NEWS_SOURCE"
AGENTS = 500
PERIODS = 400
METRICS = [
    "fake_repost_share_all", "fake_repost_share_final100",
    "target_share_all", "target_share_final100",
    "fake_repost_share_decisions", "target_share_decisions", "participation_rate",
    "cumulative_fake_reposts", "target_unique_reposters_p400",
]
PATH_COLUMNS = ["input_workbook", "result_workbook", "log"]
RUN_ID_COLUMNS = ["research_question", "condition", "seed"]
EXPECTED_RUNS = {"RQ1": 180, "RQ2": 570, "RQ3": 900, "RQ4": 360}
EXPECTED_CONDITIONS = {"RQ1": 6, "RQ2": 19, "RQ3": 30, "RQ4": 24}

warnings.filterwarnings("ignore", message="Workbook contains no default style")


def _sheet_rows(workbook, name: str) -> tuple[list[str], list[tuple]]:
    sheet = workbook[name]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    return headers, list(rows)


def _read_run(record: dict[str, object]) -> dict[str, object]:
    path = Path(str(record["result_workbook"]))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rh, repost_rows = _sheet_rows(workbook, "RepostsPerSource")
        fh, fake_rows = _sheet_rows(workbook, "FakeNewsPerSource")
        uh, unique_rows = _sheet_rows(workbook, "UniqueRepostersPerSource")
    finally:
        workbook.close()
    if not (len(repost_rows) == len(fake_rows) == len(unique_rows) == PERIODS):
        raise ValueError(f"{path}: expected {PERIODS} rows in each analytical sheet")
    rix, fix, uix = ({name: i for i, name in enumerate(h)} for h in (rh, fh, uh))
    if "Simulation" in fix:
        fix["SimulationId"] = fix["Simulation"]
    periods = [int(row[rix["Period"]]) for row in repost_rows]
    if periods != list(range(1, PERIODS + 1)):
        raise ValueError(f"{path}: periods are incomplete, duplicated, or out of order")

    fake_counts, target_counts, target_fake, decision_counts = [], [], [], []
    source_counts = {source: [] for source in SOURCES}
    period_metrics = []
    for row_index, (rr, fr, ur) in enumerate(zip(repost_rows, fake_rows, unique_rows), start=1):
        counts = [int(rr[rix[source]]) for source in SOURCES]
        states = [int(fr[fix[source]]) for source in SOURCES]
        if any(state not in (0, 1) for state in states):
            raise ValueError(f"{path}: fake-news states are not binary")
        total = sum(counts)
        reach_limited = pd.notna(record.get("target_reach"))
        activity = record.get("user_activity", 1.0)
        activity_limited = pd.notna(activity) and float(activity) < 1.0
        if total < 0 or total > AGENTS or (not reach_limited and not activity_limited and total != AGENTS):
            raise ValueError(f"{path}: invalid repost total {total}")
        fake_counts.append(sum(count * state for count, state in zip(counts, states)))
        target_counts.append(counts[SOURCES.index(TARGET)])
        target_fake.append(states[SOURCES.index(TARGET)])
        decision_counts.append(total)
        for source, count in zip(SOURCES, counts):
            source_counts[source].append(count)
        period_metrics.append({
            "research_question": record["research_question"], "condition": record["condition"],
            "seed": record["seed"], "period": row_index, "total_decisions": total,
            "false_reposts": fake_counts[-1], "target_selections": target_counts[-1],
            "target_unique_reposters": int(ur[uix[TARGET]]),
            **{f"selections_{source.lower()}": count for source, count in zip(SOURCES, counts)},
        })

    start = int(record["start_period"]) if pd.notna(record.get("start_period")) else 1
    raw_end = int(record["end_period"]) if pd.notna(record.get("end_period")) else -1
    end = PERIODS if raw_end < 0 else min(raw_end, PERIODS)
    during = slice(start - 1, end)
    post_start = end if raw_end >= 0 else PERIODS
    post_end = min(post_start + 100, PERIODS)
    post_values = fake_counts[post_start:post_end]
    result = dict(record)
    total_decisions = int(sum(decision_counts))
    total_fake = int(sum(fake_counts))
    total_target = int(sum(target_counts))
    final100_decisions = int(sum(decision_counts[-100:]))
    result.update({
        "fake_repost_share_all": float(np.mean(fake_counts) / AGENTS),
        "fake_repost_share_final100": float(np.mean(fake_counts[-100:]) / AGENTS),
        "target_share_all": float(np.mean(target_counts) / AGENTS),
        "target_share_final100": float(np.mean(target_counts[-100:]) / AGENTS),
        "fake_repost_share_decisions": total_fake / total_decisions if total_decisions else math.nan,
        "target_share_decisions": total_target / total_decisions if total_decisions else math.nan,
        "participation_rate": total_decisions / (AGENTS * PERIODS),
        "fake_repost_share_decisions_final100": sum(fake_counts[-100:]) / final100_decisions if final100_decisions else math.nan,
        "target_share_decisions_final100": sum(target_counts[-100:]) / final100_decisions if final100_decisions else math.nan,
        "total_decisions": total_decisions,
        "cumulative_fake_reposts": total_fake,
        "target_fake_rate_all": float(np.mean(target_fake)),
        "target_unique_reposters_p400": int(unique_rows[-1][uix[TARGET]]),
        "target_unique_reposters_p25": int(unique_rows[24][uix[TARGET]]),
        "target_unique_reposters_p50": int(unique_rows[49][uix[TARGET]]),
        "target_unique_reposters_p100": int(unique_rows[99][uix[TARGET]]),
        "target_unique_reposters_p200": int(unique_rows[199][uix[TARGET]]),
        "target_first_repost_period": next((i + 1 for i, value in enumerate(target_counts) if value > 0), math.nan),
        "fake_repost_share_during": float(np.mean(fake_counts[during]) / AGENTS),
        "fake_repost_share_post100": float(np.mean(post_values) / AGENTS) if post_values else math.nan,
        **{f"selections_{source.lower()}_all": int(sum(values)) for source, values in source_counts.items()},
        "_period_metrics": period_metrics,
    })
    final_unique = result["target_unique_reposters_p400"]
    for fraction in (0.25, 0.50, 0.75):
        threshold = final_unique * fraction
        result[f"target_time_to_{int(fraction * 100)}pct_final"] = next(
            (index + 1 for index, row in enumerate(unique_rows) if int(row[uix[TARGET]]) >= threshold), math.nan)
    return result


def paired_stats(treatment: pd.DataFrame, control: pd.DataFrame, metric: str) -> dict[str, float | int]:
    paired = treatment[["seed", metric]].merge(
        control[["seed", metric]], on="seed", suffixes=("_t", "_c"), validate="one_to_one")
    diff = paired[f"{metric}_t"].astype(float) - paired[f"{metric}_c"].astype(float)
    n = len(diff)
    if n < 2:
        raise ValueError(f"At least two paired seeds are required for {metric}")
    mean, sd = float(diff.mean()), float(diff.std(ddof=1))
    se = sd / math.sqrt(n)
    critical = float(stats.t.ppf(0.975, n - 1))
    test = stats.ttest_rel(paired[f"{metric}_t"], paired[f"{metric}_c"])
    return {
        "n_pairs": n, "effect": mean, "se": se,
        "ci95_low": mean - critical * se, "ci95_high": mean + critical * se,
        "cohens_dz": mean / sd if sd else math.nan, "p_value": float(test.pvalue),
    }


def _control_for(runs: pd.DataFrame, row: pd.Series) -> pd.DataFrame:
    rq = row.research_question
    control = runs[(runs.research_question == rq) & runs.condition.str.startswith("control")]
    if rq == "RQ3":
        control = control[(control.target_reach == row.target_reach) & (control.wom == row.wom)]
    elif rq == "RQ4":
        control = control[(control.contacts == row.contacts) &
                          (control.memory == row.memory) &
                          (control.memory_half_life == row.memory_half_life)]
    return control


def contrasts(runs: pd.DataFrame) -> pd.DataFrame:
    records = []
    treatment = runs[~runs.condition.str.startswith("control")]
    for (rq, condition), group in treatment.groupby(["research_question", "condition"], sort=False):
        first = group.iloc[0]
        control = _control_for(runs, first)
        if control.empty:
            raise ValueError(f"No matched control for {rq}/{condition}")
        for metric in METRICS:
            record = {"research_question": rq, "condition": condition, "metric": metric,
                      "strategy": first.strategy, "start_period": first.start_period,
                      "end_period": first.end_period, "target_reach": first.target_reach,
                      "wom": first.wom, "contacts": first.contacts, "memory": first.memory,
                      "memory_half_life": first.memory_half_life}
            record.update(paired_stats(group, control, metric))
            records.append(record)
    result = pd.DataFrame(records)
    result["p_holm"] = np.nan
    for (_, metric), index in result.groupby(["research_question", "metric"]).groups.items():
        ordered = result.loc[index, "p_value"].sort_values()
        adjusted = np.maximum.accumulate(ordered.to_numpy() * (len(ordered) - np.arange(len(ordered))))
        result.loc[ordered.index, "p_holm"] = np.minimum(adjusted, 1.0)
    return result


def condition_summary(runs: pd.DataFrame) -> pd.DataFrame:
    keys = ["research_question", "condition", "strategy", "start_period", "end_period",
            "target_reach", "wom", "contacts", "memory", "memory_half_life"]
    rows = []
    for values, group in runs.groupby(keys, dropna=False, sort=False):
        row = dict(zip(keys, values)); row["n"] = len(group)
        extra = ["fake_repost_share_during", "fake_repost_share_post100", "total_decisions",
                 "target_unique_reposters_p25", "target_unique_reposters_p50",
                 "target_unique_reposters_p100", "target_unique_reposters_p200",
                 "target_time_to_25pct_final", "target_time_to_50pct_final", "target_time_to_75pct_final"]
        extra += [f"selections_{source.lower()}_all" for source in SOURCES]
        for metric in METRICS + extra:
            data = group[metric].dropna().astype(float)
            row[f"{metric}_mean"] = data.mean() if len(data) else math.nan
            row[f"{metric}_sd"] = data.std(ddof=1) if len(data) > 1 else math.nan
        rows.append(row)
    return pd.DataFrame(rows)


def public_runs(runs: pd.DataFrame) -> pd.DataFrame:
    """Return the publication-safe, one-row-per-run analytical dataset."""
    return runs.drop(columns=PATH_COLUMNS, errors="ignore").copy()


def validate_runs(runs: pd.DataFrame) -> dict[str, object]:
    required = set(RUN_ID_COLUMNS + METRICS + ["status"])
    missing = sorted(required.difference(runs.columns))
    if missing:
        raise ValueError(f"Run dataset is missing required columns: {', '.join(missing)}")
    if len(runs) != sum(EXPECTED_RUNS.values()):
        raise ValueError(f"Expected 2,010 runs; found {len(runs)}")
    if runs.duplicated(RUN_ID_COLUMNS).any():
        raise ValueError("Run dataset contains duplicated research-question/condition/seed rows")
    if set(runs["status"]) != {"COMPLETE"}:
        raise ValueError("Every published run must have COMPLETE status")
    actual_runs = runs.groupby("research_question").size().to_dict()
    actual_conditions = runs.groupby("research_question").condition.nunique().to_dict()
    if actual_runs != EXPECTED_RUNS:
        raise ValueError(f"Unexpected run counts: {actual_runs}")
    if actual_conditions != EXPECTED_CONDITIONS:
        raise ValueError(f"Unexpected condition counts: {actual_conditions}")
    for metric in ["fake_repost_share_all", "fake_repost_share_final100",
                   "target_share_all", "target_share_final100", "participation_rate",
                   "fake_repost_share_decisions", "target_share_decisions"]:
        if not runs[metric].between(0, 1).all():
            raise ValueError(f"{metric} contains values outside [0, 1]")
    if not runs["target_unique_reposters_p400"].between(0, AGENTS).all():
        raise ValueError("target_unique_reposters_p400 contains an invalid count")
    return {
        "status": "PASS", "workbooks": len(runs), "agents": AGENTS, "periods": PERIODS,
        "runs_by_rq": actual_runs, "conditions_by_rq": actual_conditions,
        "method": "paired by seed; Student-t 95% confidence intervals; Holm adjustment within RQ and metric",
    }


def rq2_window_effects(periods: pd.DataFrame, runs: pd.DataFrame) -> pd.DataFrame:
    """Paired RQ2 effects in matched campaign and post-removal windows."""
    if periods.empty:
        return pd.DataFrame()
    metadata = runs[runs.research_question == "RQ2"].set_index(["condition", "seed"])
    control = periods[(periods.research_question == "RQ2") & (periods.condition == "control")]
    rows = []
    treatments = runs[(runs.research_question == "RQ2") & ~runs.condition.str.startswith("control")]
    for condition, group in treatments.groupby("condition", sort=False):
        first = group.iloc[0]
        start = int(first.start_period)
        end = PERIODS if int(first.end_period) < 0 else int(first.end_period)
        windows = [("during", start, end)]
        if int(first.end_period) >= 0:
            windows += [("post_1_25", end + 1, min(end + 25, PERIODS)),
                        ("post_26_50", end + 26, min(end + 50, PERIODS)),
                        ("post_51_100", end + 51, min(end + 100, PERIODS))]
        for window, low, high in windows:
            if low > high: continue
            treatment_periods = periods[(periods.condition == condition) & periods.period.between(low, high)]
            control_periods = control[control.period.between(low, high)]
            paired_rows = []
            for seed in group.seed:
                t = treatment_periods[treatment_periods.seed == seed]
                c = control_periods[control_periods.seed == seed]
                if t.empty or c.empty: continue
                def metrics(frame):
                    decisions = frame.total_decisions.sum()
                    return {
                        "false_rate_agent_period": frame.false_reposts.sum() / (AGENTS * len(frame)),
                        "false_share_decisions": frame.false_reposts.sum() / decisions if decisions else math.nan,
                        "target_share_decisions": frame.target_selections.sum() / decisions if decisions else math.nan,
                        "participation_rate": decisions / (AGENTS * len(frame)),
                    }
                tm, cm = metrics(t), metrics(c)
                paired_rows.append({"seed": seed, **{f"{key}_t": value for key, value in tm.items()},
                                    **{f"{key}_c": value for key, value in cm.items()}})
            paired = pd.DataFrame(paired_rows)
            for metric in ("false_rate_agent_period", "false_share_decisions",
                           "target_share_decisions", "participation_rate"):
                if len(paired) < 2: continue
                stats_row = paired_stats(
                    paired[["seed", f"{metric}_t"]].rename(columns={f"{metric}_t": metric}),
                    paired[["seed", f"{metric}_c"]].rename(columns={f"{metric}_c": metric}), metric)
                rows.append({"condition": condition, "strategy": first.strategy,
                             "start_period": start, "end_period": int(first.end_period),
                             "window": window, "window_start": low, "window_end": high,
                             "metric": metric, **stats_row})
    return pd.DataFrame(rows)


def write_checksums(output: Path) -> None:
    candidates = [path for path in output.rglob("*")
                  if path.is_file() and path.name != "SHA256SUMS"]
    lines = []
    for path in sorted(candidates):
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        lines.append(f"{digest}  {path.relative_to(output).as_posix()}")
    (output / "SHA256SUMS").write_text("\n".join(lines) + "\n", encoding="utf-8")


def _primary(effects: pd.DataFrame, rq: str) -> pd.DataFrame:
    return effects[(effects.research_question == rq) &
                   (effects.metric == "fake_repost_share_decisions")].copy()


def save_figures(effects: pd.DataFrame, output: Path) -> None:
    figures = output / "figures"; figures.mkdir(parents=True, exist_ok=True)
    plt.rcParams.update({"font.size": 10, "axes.labelsize": 10,
                          "axes.titlesize": 11, "figure.dpi": 150})
    q1 = _primary(effects, "RQ1").sort_values("effect")
    fig, ax = plt.subplots(figsize=(8.6, 4.4)); y = np.arange(len(q1))
    ax.errorbar(q1.effect * 100, y, xerr=[(q1.effect-q1.ci95_low)*100, (q1.ci95_high-q1.effect)*100], fmt="o", color="#9c2f2f", capsize=3)
    ax.axvline(0, color="black", lw=.7); ax.set_yticks(y, q1.condition.str.replace("_", " "))
    ax.set_xlabel("Treatment--control difference (percentage points)"); ax.set_title("RQ1: imitation strategies versus no strategy")
    fig.tight_layout(); fig.savefig(figures / "rq1-strategy-effects.svg"); fig.savefig(figures / "rq1-strategy-effects.png", dpi=300); plt.close(fig)

    q2 = _primary(effects, "RQ2").copy(); q2["strategy_label"] = q2.condition.str.split("_p").str[0]
    q2["start"] = q2.start_period.astype(int); q2["duration"] = q2.apply(lambda r: "Permanent" if int(r.end_period) < 0 else str(int(r.end_period-r.start_period+1)), axis=1)
    fig, axes = plt.subplots(1, 2, figsize=(8.4, 3.8), sharey=True)
    for ax, (strategy, group) in zip(axes, q2.groupby("strategy_label")):
        for duration, part in group.groupby("duration"):
            part = part.sort_values("start"); ax.plot(part.start, part.effect*100, marker="o", label=duration)
        ax.axhline(0,color="black",lw=.7); ax.set_title(strategy.title()); ax.set_xlabel("Campaign start period")
    axes[0].set_ylabel("Paired effect (percentage points)"); axes[1].legend(title="Duration")
    fig.suptitle("RQ2: timing and duration of malicious-source strategies"); fig.tight_layout(); fig.savefig(figures / "rq2-timing-duration.svg"); fig.savefig(figures / "rq2-timing-duration.png",dpi=300); plt.close(fig)

    q3 = _primary(effects, "RQ3").copy(); q3["strategy_label"] = q3.condition.str.split("_reach").str[0]
    fig, axes = plt.subplots(1, 2, figsize=(8.4, 3.8), sharey=True)
    for ax, (wom, group) in zip(axes, q3.groupby("wom")):
        for strategy, part in group.groupby("strategy_label"):
            part = part.sort_values("target_reach"); ax.plot(part.target_reach, part.effect*100, marker="o", label=strategy.title())
        ax.axhline(0,color="black",lw=.7); ax.set_title("Social recommendations " + ("on" if wom else "off")); ax.set_xlabel("Malicious-source reach (%)")
    axes[0].set_ylabel("Strategy effect (percentage points)"); axes[1].legend()
    fig.suptitle("RQ3: reach restrictions and social recommendations"); fig.tight_layout(); fig.savefig(figures / "rq3-reach-mitigation.svg"); fig.savefig(figures / "rq3-reach-mitigation.png",dpi=300); plt.close(fig)

    q4 = _primary(effects, "RQ4").sort_values(["contacts","memory_half_life"]); labels = q4.apply(lambda r: f"contacts={int(r.contacts)}, " + ("window=25" if int(r.memory)==25 else f"half-life={r.memory_half_life:g}"),axis=1)
    fig, ax = plt.subplots(figsize=(7.2, 5)); y=np.arange(len(q4))
    ax.errorbar(q4.effect*100,y,xerr=[(q4.effect-q4.ci95_low)*100,(q4.ci95_high-q4.effect)*100],fmt="o",color="#315b7d",capsize=3)
    ax.axvline(0,color="black",lw=.7); ax.set_yticks(y,labels); ax.set_xlabel("Credibility-strategy effect (percentage points)"); ax.set_title("RQ4: sensitivity to memory and connectivity")
    fig.tight_layout(); fig.savefig(figures / "rq4-sensitivity.svg"); fig.savefig(figures / "rq4-sensitivity.png",dpi=300); plt.close(fig)


def save_rq2_window_figure(window_effects: pd.DataFrame, output: Path) -> None:
    """Forest plots keep categorical campaign windows separate and show uncertainty."""
    if window_effects.empty:
        return
    figures = output / "figures"; figures.mkdir(parents=True, exist_ok=True)
    data = window_effects[window_effects.window.isin(["during", "post_1_25"]) &
                          window_effects.metric.isin(["false_share_decisions", "target_share_decisions"])].copy()
    conditions = list(dict.fromkeys(data.condition))
    def descriptive_label(value: str) -> str:
        strategy, start, duration = value.split("_")
        strategy = "Informational" if strategy == "informational" else "Credibility"
        start = start.removeprefix("p")
        duration = "through period 400" if duration == "permanent" else duration.removeprefix("d") + " periods"
        return f"{strategy}; start {start}; {duration}"
    labels = [descriptive_label(value) for value in conditions]
    fig, axes = plt.subplots(1, 2, figsize=(12.8, 8.0), sharey=True)
    colors = {"during": "#2b6f92", "post_1_25": "#b3532f"}
    offsets = {"during": -0.13, "post_1_25": 0.13}
    for ax, metric, title in zip(axes,
            ["target_share_decisions", "false_share_decisions"],
            ["Experimental-source share", "False share among decisions"]):
        part = data[data.metric == metric]
        for window in ("during", "post_1_25"):
            values = part[part.window == window].set_index("condition").reindex(conditions)
            valid = values.effect.notna()
            y = np.arange(len(conditions))[valid] + offsets[window]
            effect = values.loc[valid, "effect"].to_numpy() * 100
            low = values.loc[valid, "ci95_low"].to_numpy() * 100
            high = values.loc[valid, "ci95_high"].to_numpy() * 100
            ax.errorbar(effect, y, xerr=[effect-low, high-effect], fmt="o", ms=3.5,
                        color=colors[window], capsize=2,
                        label="During campaign" if window == "during" else "Periods 1--25 after removal")
        ax.axvline(0, color="black", lw=.7); ax.set_title(title)
        ax.set_xlabel("Treatment--control difference (percentage points)")
    axes[0].set_yticks(np.arange(len(conditions)), labels)
    axes[1].legend(loc="lower right", fontsize=9)
    fig.suptitle("RQ2 matched campaign and immediate post-removal windows")
    fig.tight_layout(); fig.savefig(figures / "rq2-matched-windows.svg")
    fig.savefig(figures / "rq2-matched-windows.png", dpi=300); plt.close(fig)


def latex_tables(summary: pd.DataFrame, effects: pd.DataFrame, output: Path) -> None:
    tables = output / "tables"; tables.mkdir(parents=True, exist_ok=True)
    for rq in ("RQ1", "RQ2", "RQ3", "RQ4"):
        data = _primary(effects, rq)[["condition","n_pairs","effect","ci95_low","ci95_high","cohens_dz","p_holm"]].copy()
        for c in ["effect","ci95_low","ci95_high"]: data[c] = data[c] * 100
        data.columns = ["Condition","n","Effect (pp)","95\\% CI low","95\\% CI high","$d_z$","Holm $p$"]
        (tables / f"{rq.lower()}-effects.tex").write_text(data.to_latex(index=False,float_format="%.3f",escape=False),encoding="utf-8")
    overview = summary.groupby("research_question",as_index=False).agg(conditions=("condition","nunique"),runs=("n","sum"))
    overview.columns=["Research question","Conditions","Runs"]
    (tables/"study-overview.tex").write_text(overview.to_latex(index=False,escape=False),encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("root", nargs="?", type=Path,
                        help="Directory containing RQ*/manifest.tsv and raw workbooks")
    parser.add_argument("--processed", type=Path,
                        help="Publication-safe run-metrics.csv; rebuild statistics without raw workbooks")
    parser.add_argument("--period-processed", type=Path,
                        help="Optional publication-safe period-metrics.csv.gz for matched-window analyses")
    parser.add_argument("--output", type=Path, default=Path("analysis/final-study"))
    parser.add_argument("--jobs", type=int, default=6)
    args = parser.parse_args()
    if (args.root is None) == (args.processed is None):
        parser.error("provide exactly one raw root or --processed run-metrics.csv")

    if args.processed is not None:
        runs = pd.read_csv(args.processed)
        periods = pd.read_csv(args.period_processed) if args.period_processed else pd.DataFrame()
    else:
        manifests = []
        for path in sorted(args.root.glob("RQ*/manifest.tsv")):
            data = pd.read_csv(path, sep="\t")
            data = data[data.status == "COMPLETE"].copy()
            manifests.append(data)
        if not manifests:
            raise ValueError(f"No RQ*/manifest.tsv files found below {args.root}")
        manifest = pd.concat(manifests, ignore_index=True)
        if len(manifest) != 2010 or manifest.result_workbook.duplicated().any():
            raise ValueError(f"Expected 2,010 unique completed runs; found {len(manifest)}")
        with ProcessPoolExecutor(max_workers=args.jobs) as pool:
            rows = list(pool.map(_read_run, manifest.to_dict("records"), chunksize=4))
        period_rows = []
        for row in rows:
            period_rows.extend(row.pop("_period_metrics"))
        runs = pd.DataFrame(rows)
        periods = pd.DataFrame(period_rows)

    runs = public_runs(runs)
    audit = validate_runs(runs)
    summary = condition_summary(runs)
    effects = contrasts(runs)
    window_effects = rq2_window_effects(periods, runs)
    args.output.mkdir(parents=True, exist_ok=True)
    runs.to_csv(args.output / "run-metrics.csv", index=False)
    if not periods.empty:
        periods.to_csv(args.output / "period-metrics.csv.gz", index=False, compression="gzip")
    manifest_columns = [column for column in runs.columns if column not in METRICS and
                        not column.startswith("fake_repost_share_") and
                        column not in {"target_fake_rate_all"}]
    runs[manifest_columns].to_csv(args.output / "run-manifest.csv", index=False)
    summary.to_csv(args.output / "condition-summary.csv", index=False)
    effects.to_csv(args.output / "paired-effects.csv", index=False)
    window_effects.to_csv(args.output / "rq2-window-effects.csv", index=False)
    save_figures(effects,args.output); save_rq2_window_figure(window_effects, args.output)
    latex_tables(summary,effects,args.output)
    (args.output/"validation.json").write_text(json.dumps(audit,indent=2),encoding="utf-8")
    write_checksums(args.output)
    print(json.dumps(audit,indent=2))


if __name__=="__main__": main()
