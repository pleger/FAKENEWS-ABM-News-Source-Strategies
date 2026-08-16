#!/usr/bin/env python3
"""Focused tests for dissemination-specific workbook transformations."""

from __future__ import annotations

import unittest

from openpyxl import Workbook, load_workbook

from prepare_workbook import (
    CREDIBILITY_ATTRIBUTE,
    NON_CREDIBILITY_ATTRIBUTES,
    is_header_scenario,
    parse_strategy_names,
    set_scenario_definition,
    set_scenario_attributes,
    set_source_reach,
)


class PrepareWorkbookTest(unittest.TestCase):
    def setUp(self) -> None:
        self.workbook = load_workbook("input/FAKENEWS_BASELINE_2.xlsx")

    def scenario_values(self) -> list[str]:
        return [
            cell.value for cell in self.workbook["Scenario"][1][3:]
            if cell.value is not None and str(cell.value).strip()
        ]

    def test_credibility_scope_copies_only_credibility(self) -> None:
        set_scenario_attributes(self.workbook["Scenario"], "credibility")
        self.assertEqual([CREDIBILITY_ATTRIBUTE], self.scenario_values())

    def test_non_credibility_scope_preserves_fake_news_probability_dimension(self) -> None:
        set_scenario_attributes(self.workbook["Scenario"], "non-credibility")
        self.assertEqual(list(NON_CREDIBILITY_ATTRIBUTES), self.scenario_values())
        self.assertNotIn(CREDIBILITY_ATTRIBUTE, self.scenario_values())

    def test_all_scope_uses_empty_attribute_shortcut(self) -> None:
        set_scenario_attributes(self.workbook["Scenario"], "all")
        self.assertEqual([], self.scenario_values())

    def test_target_source_reach_is_replaced(self) -> None:
        set_source_reach(self.workbook["SourceReach"], "FAKE_NEWS_SOURCE", 46.7)
        values = {row[0].value: row[1].value for row in self.workbook["SourceReach"].iter_rows()}
        self.assertEqual(46.7, values["FAKE_NEWS_SOURCE"])

    def test_header_scenario_writes_strategy_and_end_period(self) -> None:
        workbook = Workbook()
        scenario = workbook.active
        scenario.title = "Scenario"
        scenario.append(["FROM", "TO", "START_PERIOD", "END_PERIOD", "STRATEGIES", "ATTRIBUTES"])
        scenario.append(["TRADITIONAL_MEDIA", "FAKE_NEWS_SOURCE", 1, -1, None, None])

        self.assertTrue(is_header_scenario(scenario))
        names = parse_strategy_names("engagement, proximity")
        set_scenario_definition(scenario, 10, 20, names, "credibility")

        self.assertEqual(10, scenario.cell(2, 3).value)
        self.assertEqual(20, scenario.cell(2, 4).value)
        self.assertEqual("ENGAGEMENT,PROXIMITY", scenario.cell(2, 5).value)
        self.assertEqual(CREDIBILITY_ATTRIBUTE, scenario.cell(2, 6).value)

    def test_legacy_scenario_resolves_named_strategy_to_attributes(self) -> None:
        names = parse_strategy_names("credibility_camouflage")
        set_scenario_definition(self.workbook["Scenario"], 15, -1, names, "all")
        self.assertEqual([CREDIBILITY_ATTRIBUTE], self.scenario_values())


if __name__ == "__main__":
    unittest.main()
