#!/usr/bin/env python3
"""Create one isolated experiment workbook without changing the source workbook."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


CONFIGURATION = {
    "PERIODS": 400,
    "AGENTS": 400,
    "REPETITIONS": 10,
    "GUI": 0,
    "LEARNING_PERIODS": 0,
    "SAVED_ENDORSEMENTS": 0,
    "SAVED_REPOSTS_PER_SOURCE": 1,
    "SAVED_DETAILED_AGENT_DECISIONS": 0,
    "SAVED_AGENT_DECISIONS": 0,
    "SAVED_FAKENEWS": 1,
    "COMPRESSED_RESULTS": 0,
}

NON_CREDIBILITY_ATTRIBUTES = (
    "Emociones evocadas positivas",
    "Emociones evocadas negativas",
    "Lenguaje sencillo",
    "Proximidad política",
    "Proximidad geográfica al problema",
    "Proximidad social al problema",
    "Sensacionalismo de la noticia",
    "Calidad de la información",
    "Diversión de la información",
    "Contenido audiovisual",
    "Enlaces",
    "Hashtags",
)

CREDIBILITY_ATTRIBUTE = "Credibilidad de la fuente"

SCENARIO_ATTRIBUTE_SELECTIONS = {
    "all": (),
    "credibility": (CREDIBILITY_ATTRIBUTE,),
    "non-credibility": NON_CREDIBILITY_ATTRIBUTES,
    # Backward-compatible name used by the earlier research-question suite.
    "engagement": NON_CREDIBILITY_ATTRIBUTES,
}

NAMED_STRATEGIES = {
    "ENGAGEMENT": (
        "Emociones evocadas positivas",
        "Emociones evocadas negativas",
        "Lenguaje sencillo",
        "Sensacionalismo de la noticia",
        "Diversión de la información",
        "Contenido audiovisual",
        "Hashtags",
    ),
    "PROXIMITY": (
        "Proximidad política",
        "Proximidad geográfica al problema",
        "Proximidad social al problema",
    ),
    "INFORMATIONAL_CAMOUFLAGE": (
        "Calidad de la información",
        "Enlaces",
    ),
    "CREDIBILITY_CAMOUFLAGE": (CREDIBILITY_ATTRIBUTE,),
}


def set_configuration(worksheet, key: str, value: int | float) -> None:
    for row in worksheet.iter_rows():
        if str(row[0].value).strip().upper() == key:
            row[1].value = value
            return
    worksheet.append([key, value])


def set_user_weight(worksheet, name: str, value: float) -> None:
    wanted = name.strip().upper()
    for row in worksheet.iter_rows():
        if str(row[0].value).strip().upper() == wanted:
            row[1].value = value
            return
    raise ValueError(f"SNSUsers does not contain attribute: {name}")


def set_source_reach(worksheet, source_name: str, percentage: float) -> None:
    """Replace one source's visibility percentage in the SourceReach worksheet."""
    wanted = source_name.strip().upper()
    for row in worksheet.iter_rows():
        if str(row[0].value).strip().upper() == wanted:
            row[1].value = percentage
            return
    raise ValueError(f"SourceReach does not contain source: {source_name}")


def set_scenario_attributes(worksheet, selection: str) -> None:
    names = SCENARIO_ATTRIBUTE_SELECTIONS[selection]
    maximum_column = max(worksheet.max_column, 3 + len(names))
    for column in range(4, maximum_column + 1):
        worksheet.cell(1, column).value = None
    for column, name in enumerate(names, start=4):
        worksheet.cell(1, column).value = name


def parse_strategy_names(value: str | None) -> tuple[str, ...]:
    if not value:
        return ()
    result = tuple(name.strip().upper() for name in value.split(",") if name.strip())
    unknown = [name for name in result if name not in NAMED_STRATEGIES]
    if unknown:
        raise ValueError(f"unknown scenario strategies: {', '.join(unknown)}")
    return result


def is_header_scenario(worksheet) -> bool:
    return (
        str(worksheet.cell(1, 1).value).strip().upper() == "FROM"
        and str(worksheet.cell(1, 3).value).strip().upper() == "START_PERIOD"
    )


def set_scenario_definition(
        worksheet, start: int, end: int, strategy_names: tuple[str, ...], selection: str) -> None:
    explicit = SCENARIO_ATTRIBUTE_SELECTIONS[selection]
    if is_header_scenario(worksheet):
        worksheet.cell(2, 3).value = start
        worksheet.cell(2, 4).value = end
        worksheet.cell(2, 5).value = ",".join(strategy_names) or None
        maximum_column = max(worksheet.max_column, 5 + len(explicit))
        for column in range(6, maximum_column + 1):
            worksheet.cell(2, column).value = None
        for column, name in enumerate(explicit, start=6):
            worksheet.cell(2, column).value = name
        return

    # Legacy workbooks have no strategy reference column. Resolve names to their concrete
    # attributes so the old Scenario schema and runtime semantics remain usable.
    resolved = []
    for strategy_name in strategy_names:
        resolved.extend(NAMED_STRATEGIES[strategy_name])
    resolved.extend(explicit)
    resolved = list(dict.fromkeys(resolved))
    worksheet.cell(1, 3).value = start
    maximum_column = max(worksheet.max_column, 3 + len(resolved))
    for column in range(4, maximum_column + 1):
        worksheet.cell(1, column).value = None
    for column, name in enumerate(resolved, start=4):
        worksheet.cell(1, column).value = name


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("destination", type=Path)
    parser.add_argument("--memory", type=int, required=True)
    parser.add_argument("--memory-half-life", type=float)
    parser.add_argument("--wom", type=int, choices=(0, 1), required=True)
    parser.add_argument("--scenario-period", type=int)
    parser.add_argument("--scenario-end-period", type=int, default=-1)
    parser.add_argument("--scenario-strategies",
                        help="Comma-separated strategy names defined by the experiment schema.")
    parser.add_argument(
        "--scenario-attributes",
        choices=tuple(SCENARIO_ATTRIBUTE_SELECTIONS),
        default="all",
    )
    parser.add_argument("--contacts", type=int)
    parser.add_argument("--friends", type=float)
    parser.add_argument("--source-reach", type=int, choices=(0, 1))
    parser.add_argument("--target-source", default="FAKE_NEWS_SOURCE")
    parser.add_argument("--target-source-reach", type=float,
                        help="Target source visibility percentage in SourceReach (0..100).")
    parser.add_argument("--base", type=float)
    parser.add_argument("--wom-weight", type=float)
    parser.add_argument("--wom-receiver-scale", type=float)
    parser.add_argument("--wom-fake-news-effect", type=int, choices=(-1, 0, 1))
    parser.add_argument("--wom-true-news-effect", type=int, choices=(-1, 0, 1))
    parser.add_argument("--periods", type=int, default=400)
    parser.add_argument("--agents", type=int, default=400)
    parser.add_argument("--repetitions", type=int, default=10)
    args = parser.parse_args()

    if args.memory < 0 and args.memory != -1:
        parser.error("--memory must be -1 or a nonnegative integer")
    if (args.memory_half_life is not None and
            args.memory_half_life != -1 and args.memory_half_life <= 0):
        parser.error("--memory-half-life must be -1 or greater than zero")
    if args.scenario_period is not None and not 1 <= args.scenario_period <= args.periods:
        parser.error(f"--scenario-period must be within 1..{args.periods}")
    if args.scenario_period is None and args.scenario_end_period != -1:
        parser.error("--scenario-end-period requires --scenario-period")
    if (args.scenario_period is not None and args.scenario_end_period != -1 and
            not args.scenario_period <= args.scenario_end_period <= args.periods):
        parser.error("--scenario-end-period must be -1 or within scenario-period..periods")
    try:
        scenario_strategies = parse_strategy_names(args.scenario_strategies)
    except ValueError as error:
        parser.error(str(error))
    if args.contacts is not None and args.contacts < 0:
        parser.error("--contacts must be nonnegative")
    if args.friends is not None and not 0 <= args.friends <= 1:
        parser.error("--friends must be within 0..1")
    if args.base is not None and args.base <= 0:
        parser.error("--base must be greater than zero")
    if args.wom_receiver_scale is not None and args.wom_receiver_scale < 0:
        parser.error("--wom-receiver-scale must be nonnegative")
    if args.target_source_reach is not None and not 0 <= args.target_source_reach <= 100:
        parser.error("--target-source-reach must be within 0..100")
    if args.periods <= 0 or args.agents <= 0 or args.repetitions < 0:
        parser.error("periods and agents must be positive; repetitions must be nonnegative")
    if not args.source.is_file():
        parser.error(f"source workbook does not exist: {args.source}")

    workbook = load_workbook(args.source)
    configuration = workbook["Configuration"]
    values = dict(CONFIGURATION)
    values.update({
        "PERIODS": args.periods,
        "AGENTS": args.agents,
        "REPETITIONS": args.repetitions,
        "MEMORY": args.memory,
        "WOM": args.wom,
        "SCENARIO": 0 if args.scenario_period is None else -2,
    })
    optional_values = {
        "CONTACTS": args.contacts,
        "FRIENDS": args.friends,
        "SOURCE_REACH": args.source_reach,
        "BASE": args.base,
        "MEMORY_HALF_LIFE": args.memory_half_life,
        "WOM_RECEIVER_SCALE": args.wom_receiver_scale,
        "WOM_FAKE_NEWS_EFFECT": args.wom_fake_news_effect,
        "WOM_TRUE_NEWS_EFFECT": args.wom_true_news_effect,
    }
    values.update({key: value for key, value in optional_values.items() if value is not None})
    for key, value in values.items():
        set_configuration(configuration, key, value)

    if args.scenario_period is not None:
        scenario = workbook["Scenario"]
        definition_row = 2 if is_header_scenario(scenario) else 1
        if not scenario.cell(definition_row, 1).value or not scenario.cell(definition_row, 2).value:
            parser.error("Scenario must define source FROM and TO")
        set_scenario_definition(scenario, args.scenario_period, args.scenario_end_period,
                                scenario_strategies, args.scenario_attributes)

    if args.wom_weight is not None:
        set_user_weight(workbook["SNSUsers"], "WORD OF MOUTH", args.wom_weight)

    if args.target_source_reach is not None:
        set_source_reach(workbook["SourceReach"], args.target_source, args.target_source_reach)

    args.destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.destination)


if __name__ == "__main__":
    main()
